"""VAHAN file adapter: parse the reportTable exports into single-source registration rows.

Uses the small committed fixtures (fixtures/vahan/*.xlsx), never the live export.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from pipeline.adapters.vahan import VahanFileAdapter, VahanFileParseError
from pipeline.build_view import build_view
from pipeline.store.revisions import apply_revisions, current_rows

MAKER = "fixtures/vahan/maker_sample.xlsx"
FUEL = "fixtures/vahan/fuel_sample.xlsx"
META = {
    "generated_at": "2026-07-15T10:00:00+05:30",
    "source": "VAHAN",
    "snapshot_id": None,
    "notes": None,
}


def _jan(rows, powertrain: bool = False) -> dict:
    if powertrain:
        return {
            (r.company_canonical, r.powertrain.value): r.value
            for r in rows
            if r.period_date.month == 1
        }
    return {r.company_canonical: r.value for r in rows if r.period_date.month == 1}


def test_maker_parse_maps_known_buckets_unknown_and_totals() -> None:
    a = VahanFileAdapter(MAKER)
    rows = a.parse(a.fetch("x"))
    assert a.modes == ["maker"]
    # every row is single-source VAHAN registrations, all-vehicle, domestic-only
    assert {r.source.value for r in rows} == {"VAHAN"}
    assert {r.category.value for r in rows} == {"ALL"}
    assert {r.flow.value for r in rows} == {"domestic"}
    assert {r.powertrain.value for r in rows} == {"all"}
    assert {r.native_frequency.value for r in rows} == {"month"}

    jan = _jan(rows)
    assert jan["Hero MotoCorp"] == 500000
    assert jan["TVS Motor Company"] == 300000
    assert jan["Ola Electric"] == 100000  # present in VAHAN, ABSENT from SIAM
    assert jan["Others"] == 10000  # the two unmapped makers (8000 + 2000), not hard-failed
    # mapped makers + Others reconcile to the reported universe total
    total = jan["Total Reported Universe"]
    assert total == 910000
    assert sum(v for k, v in jan.items() if k != "Total Reported Universe") == total


def test_fuel_parse_counts_only_battery_electric_as_ev() -> None:
    a = VahanFileAdapter(FUEL)
    rows = a.parse(a.fetch("x"))
    assert a.modes == ["fuel"]
    jan = _jan(rows, powertrain=True)
    assert jan[("Total Reported Universe", "all")] == 910000  # every fuel
    # EV = PURE EV (120000) + ELECTRIC(BOV) (30000). STRONG HYBRID EV / DIESEL / PETROL excluded.
    assert jan[("Total Reported Universe", "ev")] == 150000


def test_batch_merges_and_dedups_the_reconciling_total() -> None:
    a = VahanFileAdapter([MAKER, FUEL])
    assert a.parse(a.fetch("x"))  # both files
    store = current_rows(apply_revisions([], a.parse(a.fetch("x"))).rows)
    jan = _jan(store, powertrain=True)
    # the identical maker-total and fuel-total collapse to ONE row (not doubled)
    assert jan[("Total Reported Universe", "all")] == 910000
    assert jan[("Total Reported Universe", "ev")] == 150000
    assert jan[("Hero MotoCorp", "all")] == 500000


def test_view_is_single_source_registrations_no_export_or_production() -> None:
    a = VahanFileAdapter([MAKER, FUEL])
    store = current_rows(apply_revisions([], a.parse(a.fetch("x"))).rows)
    view = build_view(store, META, "VAHAN")
    m = view["meta"]
    assert m["source"] == "VAHAN"
    assert m["category"] == "VAHAN"
    assert "registrations" in m["source_universe_label"].lower()
    assert view["flows"] == ["domestic"]  # never export or production
    assert m["has_production"] is False
    assert m["has_ev"] is True
    assert "Ola Electric" in view["companies"]
    # EV penetration = EV / total = 150000 / 910000 for JAN
    pen = view["ev_penetration"]["domestic"]["month"]["2026-01"]
    assert abs(pen - 150000 / 910000) < 1e-9


def test_malformed_reporttable_raises(tmp_path: Path) -> None:
    # a workbook with the right sheet name but no month header must fail (→ quarantine upstream)
    bad = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "reportTable"
    wb.active.cell(1, 1, "Maker Month Wise Data For All State (2026)")
    wb.active.cell(2, 2, "GARBAGE WITH NO MONTH ROW")
    wb.save(bad)
    a = VahanFileAdapter(str(bad))
    with pytest.raises(VahanFileParseError):
        a.parse(a.fetch("x"))
