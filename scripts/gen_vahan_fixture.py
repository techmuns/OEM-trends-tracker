"""Generate the small committed VAHAN test fixtures (never the live export).

Two tiny reportTable.xlsx files mirroring the vahan4dashboard "Maker Month Wise" and
"Fuel Month Wise" downloads: a title row carrying the year, an 'S No / Maker|Fuel' header
row, a month-header row, Indian-grouped string values, and a YTD TOTAL column. The maker and
fuel grand totals reconcile per month (as the real exports do). Includes an EV maker (Ola,
present in VAHAN / absent from SIAM), two unmapped makers (→ Others), and hybrid fuels (which
must NOT count as EV). Deterministic; re-run to regenerate.

    uv run python scripts/gen_vahan_fixture.py
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

OUT = Path("fixtures/vahan")
MONTHS = ["JAN", "FEB", "MAR"]


def _india(n: int) -> str:
    # Indian digit grouping: 12,34,567 — the exact string shape the export uses.
    s = str(abs(n))
    if len(s) <= 3:
        body = s
    else:
        head, tail = s[:-3], s[-3:]
        parts = []
        while len(head) > 2:
            parts.insert(0, head[-2:])
            head = head[:-2]
        parts.insert(0, head)
        body = ",".join(parts) + "," + tail
    return ("-" if n < 0 else "") + body


def _write(path: Path, entity_label: str, rows: dict[str, list[int]], year: int, kind: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "reportTable"
    ws.cell(1, 1, f"{kind} Month Wise Data  For All State ({year})")
    ws.cell(2, 1, "S No")
    ws.cell(2, 2, f"\xa0\xa0\xa0 {entity_label} \xa0\xa0\xa0")
    ws.cell(2, 3, "Month Wise ")
    ws.cell(2, 3 + len(MONTHS), "\xa0TOTAL\xa0")
    # row 3 blank; row 4 month headers (TOTAL column left blank there)
    for i, mo in enumerate(MONTHS):
        ws.cell(4, 3 + i, mo)
    for r, (name, vals) in enumerate(rows.items(), start=5):
        ws.cell(r, 1, r - 4)
        ws.cell(r, 2, name)
        for i, v in enumerate(vals):
            ws.cell(r, 3 + i, _india(v))
        ws.cell(r, 3 + len(MONTHS), _india(sum(vals)))
    OUT.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"[gen_vahan_fixture] wrote {path}")


MAKERS = {
    "HERO MOTOCORP LTD": [500000, 460000, 520000],
    "TVS MOTOR COMPANY LTD": [300000, 280000, 310000],
    "OLA ELECTRIC TECHNOLOGIES PVT LTD": [100000, 90000, 110000],  # EV maker, present in VAHAN
    "SUPER OBSCURE MOTORS PVT LTD": [8000, 7000, 9000],  # unmapped -> Others
    "TINY GARAGE WORKS": [2000, 3000, 1000],  # unmapped -> Others
}
# grand totals per month: 910000 / 840000 / 950000
FUELS = {
    "PETROL": [700000, 650000, 720000],
    "DIESEL": [10000, 8000, 12000],
    "PURE EV": [120000, 110000, 130000],  # EV
    "ELECTRIC(BOV)": [30000, 25000, 35000],  # EV
    "STRONG HYBRID EV": [40000, 37000, 43000],  # hybrid — NOT EV
    "CNG ONLY": [10000, 10000, 10000],
}


def main() -> int:
    _write(OUT / "maker_sample.xlsx", "Maker", MAKERS, 2026, "Maker")
    _write(OUT / "fuel_sample.xlsx", "Fuel", FUELS, 2026, "Fuel")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
