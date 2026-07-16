"""Watched-folder ingest — the automated pipeline entrypoint (monthly cron + dispatch).

Drop a source workbook in ``data/raw/incoming/`` and everything downstream runs unattended.
A single File-1 workbook carries several categories (2W, PV, …); each is a separate adapter
with its own store / snapshot / bundle / view. Every category runs independently:

    for each (category, adapter) in the file:
        parse -> apply revisions -> gates
          ├─ all pass -> snapshot -> normalized store -> per-category view (data/bundle/<cat>.json)
          └─ any fail -> quarantine the file -> keep the last good views LIVE -> non-zero exit

No human approval step; a bad file never poisons the store nor takes the dashboard down.
Idempotent: re-running the same file is a no-op. A manifest (categories.json) lists the
categories the UI can switch between.
"""

from __future__ import annotations

import json
import re
import shutil
import sys
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl

from pipeline.adapters.base import SourceAdapter
from pipeline.adapters.excel_cv import CvQuarterlyAdapter
from pipeline.adapters.excel_nested import NestedBlockAdapter
from pipeline.adapters.excel_spark import ExcelSparkAdapter
from pipeline.adapters.siam_monthly import SiamMonthlyAdapter
from pipeline.adapters.vahan import VahanFileAdapter
from pipeline.build_bundle import build_bundle
from pipeline.build_view import BUNDLE_DIR, write_manifest, write_view
from pipeline.contract.constants import FILE1_LAST_PERIOD
from pipeline.store.normalized import load_normalized, save_normalized
from pipeline.store.revisions import apply_revisions, current_rows
from pipeline.store.snapshot import snapshot_id, write_snapshot
from pipeline.validate.gates import QUARANTINE_DIR, GateContext, quarantine_path, run_gates

SOURCE = "SIAM"
INCOMING_DIR = Path("data/raw/incoming")
PROCESSED_DIR = Path("data/raw/processed")
_IST = timezone(timedelta(hours=5, minutes=30))
INGEST_TS = datetime(2026, 7, 15, 10, 0, tzinfo=_IST)

FILE1_SHEET = "OEM - Summary - 2W, PV, 3W"
FILE2_SHEET = "Two Wheelers"
VAHAN_SHEET = "reportTable"  # the vahan4dashboard "Report" download sheet

# A VAHAN export carries no category column, so the uploader stamps it into the FILENAME
# (VAHAN-2W-…, VAHAN_PV_…, or plain VAHAN…/VAHAN-ALL-… for the unfiltered all-vehicle export).
# That token routes the file to its own tab/view — 2W registrations never touch the SIAM 2W
# view, and each VAHAN category is a distinct store/view/manifest key.
VAHAN_VIEW_KEYS = {
    "ALL": "VAHAN",
    "2W": "VAHAN2W",
    "PV": "VAHANPV",
    "3W": "VAHAN3W",
    "CV": "VAHANCV",
}


def _vahan_category(path: Path) -> str:
    """The vehicle category a VAHAN file was filtered to, read from its filename token.
    (Uses a non-alphanumeric lookahead, not \\b, so 'VAHAN_PV_…' matches — '_' is a word char.)"""
    m = re.search(r"VAHAN[-_ ]?(ALL|2W|PV|3W|CV)(?![A-Za-z0-9])", path.stem, re.IGNORECASE)
    return m.group(1).upper() if m else "ALL"


def detect_adapters(path: Path, ts: datetime) -> list[tuple[str, SourceAdapter]]:
    """Return the (category, adapter) pairs for a workbook — by sheet fingerprint."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = set(wb.sheetnames)
    finally:
        wb.close()
    if FILE1_SHEET in sheets:
        adapters: list[tuple[str, SourceAdapter]] = [
            ("2W", ExcelSparkAdapter(path, ingest_date=ts)),
            ("PV", NestedBlockAdapter("PV", path, ingest_date=ts)),
            ("3W", NestedBlockAdapter("3W", path, ingest_date=ts)),
        ]
        if "CV" in sheets:
            adapters.append(("CV", CvQuarterlyAdapter(path, ingest_date=ts)))
        return adapters
    if FILE2_SHEET in sheets and "Passenger Vehicle" in sheets:
        return [("2W", SiamMonthlyAdapter(path, ingest_date=ts))]
    # VAHAN registrations export (maker-wise OR fuel-wise) — its own SEPARATE tab, never a
    # SIAM category. The adapter self-detects maker vs fuel; the filename token picks the tab.
    if VAHAN_SHEET in sheets:
        cat = _vahan_category(path)
        return [(VAHAN_VIEW_KEYS[cat], VahanFileAdapter(path, ingest_date=ts, category=cat))]
    return []


def _extras(adapter: SourceAdapter) -> dict:
    extras: dict = {}
    if isinstance(adapter, (ExcelSparkAdapter, NestedBlockAdapter, CvQuarterlyAdapter)):
        extras["reconciliation"] = adapter.reconciliation
    if isinstance(adapter, SiamMonthlyAdapter):
        extras["seam_reference"] = adapter.seam_reference
    return extras


def _notes(category: str, live, source: str = "SIAM") -> str:
    latest = max(r.period_date for r in live).isoformat()
    if source == "VAHAN":
        return (
            f"Data through {latest}. Source: VAHAN registrations (manually downloaded export "
            "from the vahan4dashboard), all-India. Registrations are a DIFFERENT measurement "
            "basis from SIAM wholesale dispatches and are never mixed with them. VAHAN reports "
            "no exports and no production."
        )
    if any("Monthly_SIAM" in r.source_file for r in live):
        return (
            f"Data through {latest}. Source: SIAM wholesale dispatches. File 1 (summary) "
            f"covers to {FILE1_LAST_PERIOD} with segment + EV detail; File 2 (monthly SIAM) "
            "extends maker-level totals past it. EV-vs-ICE and segments end "
            f"{FILE1_LAST_PERIOD}; maker values differ up to ~18% at the seam (File 1 not superseded)."
        )
    if category == "CV":
        return (
            f"Data through {latest} (CV is reported QUARTERLY, not monthly). Source: SIAM "
            "wholesale dispatches. EV is not derivable for CV — electric-bus makers sit inline "
            "among ICE makers, so EV volume/share is not shown."
        )
    if category in ("PV", "3W"):
        return (
            f"Data through {latest}. Source: SIAM wholesale dispatches. EV is not derivable "
            f"for {category} — EV-only makers sit inline among ICE makers, so EV volume/share is "
            "not shown."
        )
    return f"Data through {latest}. Source: SIAM wholesale dispatches."


def process_category(category: str, adapter: SourceAdapter, ts: datetime) -> int:
    # Source is the adapter's, not a global — SIAM and VAHAN can both flow through here. Every
    # SIAM adapter reports source_id="SIAM", so this is identical for existing categories.
    source = getattr(adapter, "source_id", None) or SOURCE
    rows = adapter.parse(adapter.fetch("as_of"))
    adapter.validate(rows)
    previous = load_normalized(category)
    outcome = apply_revisions(previous, rows)
    store_rows = outcome.rows
    live = current_rows(store_rows)

    if previous and not outcome.added_keys and not outcome.revised_keys:
        print(f"[ingest] {category}: no change (idempotent).")
        return 0

    report = run_gates(GateContext(rows=live, previous_rows=previous, extras=_extras(adapter)))
    print(f"[ingest] {category}: {report.summary()}")
    for r in report.results:
        if r.status.value != "skip":
            print(f"      {r.name}: {r.status.value} — {r.message}")
    if not report.accepted:
        print(
            f"[ingest] {category}: FAILED — "
            + "; ".join(f"{f.name}: {f.message}" for f in report.failures),
            file=sys.stderr,
        )
        return 1

    save_normalized(category, store_rows)
    write_snapshot(store_rows, ts, category=category)
    build_bundle(
        store_rows,
        generated_at=ts,
        source=source,
        category=category,
        snapshot_id=snapshot_id(ts),
        notes=_notes(category, live, source),
        bundle_path=BUNDLE_DIR / f"bundle-{category.lower()}.json",
    )
    write_view(
        store_rows,
        {
            "generated_at": ts.isoformat(),
            "source": source,
            "snapshot_id": snapshot_id(ts),
            "notes": _notes(category, live, source),
        },
        category,
    )
    print(
        f"[ingest] {category}: accepted. added={len(outcome.added_keys)} "
        f"revised={len(outcome.revised_keys)} latest={max(r.period_date for r in live)}"
    )
    return 0


def process_file(path: Path, ts: datetime = INGEST_TS) -> int:
    # A bad file must never crash the run nor poison the store: any failure (unreadable
    # workbook, parse error, unresolved maker, gate rejection) quarantines the file and
    # leaves every last-good view untouched.
    try:
        adapters = detect_adapters(path, ts)
    except Exception as e:  # noqa: BLE001 - defensive: an unreadable/corrupt workbook
        _quarantine(path, f"unreadable workbook: {type(e).__name__}: {e}")
        return 1
    if not adapters:
        _quarantine(path, "unrecognized workbook (no known category sheet)")
        return 1
    print(f"[ingest] {path.name}: categories {[c for c, _a in adapters]}")
    for category, adapter in adapters:
        try:
            failed = process_category(category, adapter, ts) != 0
        except Exception as e:  # noqa: BLE001 - parse/resolve error -> quarantine, don't crash
            _quarantine(path, f"{category}: {type(e).__name__}: {e}")
            return 1
        if failed:
            _quarantine(path, f"{category} gates failed")
            return 1
    _rebuild_manifest()
    _archive(path)
    return 0


def _rebuild_manifest() -> None:
    # Views are data/bundle/<cat>.json. Skip the manifest and every canonical bundle
    # (legacy bundle.json + per-category bundle-<cat>.json) — those are not views.
    views = []
    for p in sorted(BUNDLE_DIR.glob("*.json")):
        if p.name == "categories.json" or p.name.startswith("bundle"):
            continue
        views.append(json.loads(p.read_text()))
    if views:
        write_manifest(views)


def _archive(path: Path) -> None:
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    dest = PROCESSED_DIR / path.name
    if path.resolve() != dest.resolve():
        shutil.move(str(path), str(dest))


def _quarantine(path: Path, reason: str) -> None:
    QUARANTINE_DIR.mkdir(parents=True, exist_ok=True)
    dest = quarantine_path(str(path))
    shutil.move(str(path), str(dest))
    print(f"[ingest] QUARANTINED {path.name} -> {dest}: {reason}", file=sys.stderr)
    print("[ingest] last good views left untouched.", file=sys.stderr)


def _is_vahan(path: Path) -> bool:
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            return VAHAN_SHEET in set(wb.sheetnames)
        finally:
            wb.close()
    except Exception:  # noqa: BLE001 - unreadable -> not VAHAN; process_file will quarantine
        return False


def _process_vahan_batch(paths: list[Path], category: str, ts: datetime) -> int:
    """One VAHAN category's files (maker-wise + fuel-wise) are ONE dataset — parse and gate
    them together, so complementary files never look like a huge row delta relative to each
    other. `category` (from the filenames) routes them to their own view/store."""
    view_key = VAHAN_VIEW_KEYS[category]
    print(f"[ingest] VAHAN[{category}] -> {view_key} batch: {[p.name for p in paths]}")
    adapter = VahanFileAdapter([str(p) for p in paths], ingest_date=ts, category=category)
    try:
        failed = process_category(view_key, adapter, ts) != 0
    except Exception as e:  # noqa: BLE001 - parse/resolve error -> quarantine the whole batch
        for p in paths:
            _quarantine(p, f"{view_key}: {type(e).__name__}: {e}")
        return 1
    if failed:
        for p in paths:
            _quarantine(p, f"{view_key} gates failed")
        return 1
    _rebuild_manifest()
    for p in paths:
        _archive(p)
    return 0


def run_incoming(incoming_dir: Path = INCOMING_DIR, ts: datetime = INGEST_TS) -> int:
    files = sorted(p for p in incoming_dir.glob("*.xlsx") if p.is_file())
    if not files:
        print("[ingest] no files in incoming/ — nothing to do.")
        return 0
    vahan_files = [f for f in files if _is_vahan(f)]
    other_files = [f for f in files if f not in vahan_files]
    worst = 0
    for path in other_files:
        worst = max(worst, process_file(path, ts))
    # group VAHAN files by their filename category token; each category ingests independently
    by_cat: dict[str, list[Path]] = defaultdict(list)
    for f in vahan_files:
        by_cat[_vahan_category(f)].append(f)
    for category, paths in sorted(by_cat.items()):
        worst = max(worst, _process_vahan_batch(paths, category, ts))
    return worst


def main(argv: list[str] | None = None) -> int:
    return run_incoming()


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
