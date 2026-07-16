"""Quarterly, multi-flow-region adapter for File 1's CV sheet.

CV is unlike 2W/PV/3W in three ways, all config-driven from `dictionaries/categories.yaml`:
  * **Quarterly-native.** The source reports quarters (columns "1QFY20"…"4QFY26"), not
    months. Per the frozen contract, quarters here are the REPORTED base
    (`period_type=quarter`, `native_frequency=quarter`, `calc_status=reported`) — never
    derived from months. Annual/half-year/market-share columns in the sheet are ignored.
  * **Three flow regions on one sheet** — "Domestic Sales", "Exports", "Production" — each
    with the same four leaf segments (M&HCV Passenger/Goods, LCV Passenger/Goods).
  * **A two-level segment hierarchy.** Each region has a reported grand total ("Total CVs")
    ingested as the industry total, plus intermediate subtotals ("Total M&HCVs", "Total
    LCVs") and a maker-rollup ("… - Overall") tail that are SKIPPED so nothing double-counts.

Same non-negotiables as the other adapters: `data_only=True`, label-driven, unresolved maker
→ hard fail, `0` ≠ blank, reported totals ingested (never reconstructed by summing makers).
EV is NOT derivable — every row is `powertrain=all`.
"""

from __future__ import annotations

import datetime as dt
from datetime import datetime
from pathlib import Path

import openpyxl

from pipeline.adapters.base import RawPayload, SourceAdapter, ValidationResult
from pipeline.aggregate.periods import fiscal_quarter_start_date
from pipeline.contract.constants import INDUSTRY_TOTAL_CANONICAL
from pipeline.contract.models import ContractRow
from pipeline.dictionaries.loader import (
    UnknownCompanyError,
    load_categories,
    load_company_resolver,
)
from pipeline.validate.reconciliation import ReconciliationData, parse_file_quarter, series_key

_IST = dt.timezone(dt.timedelta(hours=5, minutes=30))


class CvParseError(RuntimeError):
    pass


class CvQuarterlyAdapter(SourceAdapter):
    source_id = "SIAM"
    native_frequency = "quarter"

    def __init__(
        self,
        source_file: str | Path,
        category: str = "CV",
        ingest_date: datetime | None = None,
        source_period: str = "4QFY26",
    ) -> None:
        self.category = category
        self.cfg = load_categories()[category]
        self.source_id = self.cfg.get("source", "SIAM")
        self.source_file = str(source_file)
        self.ingest_date = ingest_date or datetime(2026, 7, 15, 9, 0, tzinfo=_IST)
        self.source_period = source_period
        self.reconciliation = ReconciliationData()
        self.warnings: list[str] = []

    def fetch(self, period: str) -> RawPayload:
        path = Path(self.source_file)
        if not path.exists():
            raise CvParseError(f"source file not found: {path}")
        return RawPayload(self.source_id, period, self.source_file, content=path)

    def parse(self, raw: RawPayload) -> list[ContractRow]:
        wb = openpyxl.load_workbook(raw.content, data_only=True)
        try:
            sheet = self.cfg["sheet"]
            if sheet not in wb.sheetnames:
                raise CvParseError(f"sheet {sheet!r} not found")
            ws = wb[sheet]
            quarter_cols = self._read_quarter_header(ws)
            quarter_cols = self._prune_incomplete_quarters(ws, quarter_cols)
            self.reconciliation = ReconciliationData()
            rows: list[ContractRow] = []
            for flow, region in self.cfg["flow_regions"].items():
                start = self._find_label(ws, region["region_start"])
                end = (
                    self._find_label(ws, region["region_end"])
                    if region.get("region_end")
                    else ws.max_row + 1
                )
                rows += self._parse_region(ws, flow, start, end, quarter_cols)
            return rows
        finally:
            wb.close()

    def validate(self, rows: list[ContractRow]) -> ValidationResult:
        errors: list[str] = []
        warnings = list(self.warnings)
        if not rows:
            errors.append("no rows parsed")
        warnings += [
            f"negative {r.value} for {series_key(r)} {r.period_date}"
            for r in rows
            if r.value is not None and r.value < 0
        ]
        return ValidationResult(ok=not errors, errors=errors, warnings=warnings)

    # --- helpers -----------------------------------------------------------------------

    def _read_quarter_header(self, ws) -> dict[int, str]:
        """col -> fiscal-quarter label ('Q1FY20'). Only true quarter columns match; the
        sheet's annual / half-year / market-share columns are ignored."""
        header_row = self.cfg.get("header_row", 1)
        out: dict[int, str] = {}
        for col in range(2, ws.max_column + 1):
            v = ws.cell(row=header_row, column=col).value
            if isinstance(v, str) and (qk := parse_file_quarter(v)) is not None:
                q, fy = qk
                out[col] = f"Q{q}{fy}"
        if not out:
            raise CvParseError("no quarter columns found in the CV header row")
        return out

    def _prune_incomplete_quarters(self, ws, quarter_cols: dict[int, str]) -> dict[int, str]:
        """Drop quarter columns that are not yet real. The workbook computes the trailing
        quarter as annual − 9M; while the annual figure is still blank that cell is a large
        NEGATIVE number. A reported industry total can never be negative, so any quarter whose
        domestic "Total CVs" is < 0 is an unfinalized formula column and is excluded."""
        dom_start = self._find_label(ws, self.cfg["flow_regions"]["domestic"]["region_start"])
        dom_end = self._find_label(ws, self.cfg["flow_regions"]["domestic"]["region_end"])
        total_row = next(
            (
                r
                for r in range(dom_start, dom_end)
                if isinstance(ws.cell(row=r, column=1).value, str)
                and ws.cell(row=r, column=1).value.strip() == self.cfg["industry_total"]
            ),
            None,
        )
        if total_row is None:
            return quarter_cols
        kept: dict[int, str] = {}
        for col, fq in quarter_cols.items():
            v = ws.cell(row=total_row, column=col).value
            if isinstance(v, (int, float)) and v < 0:
                self.warnings.append(
                    f"CV: dropped unfinalized quarter {fq} (domestic total {v:.0f} < 0)"
                )
                continue
            kept[col] = fq
        return kept

    def _find_label(self, ws, label: str) -> int:
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if isinstance(a, str) and a.strip() == label:
                return r
        raise CvParseError(f"expected label {label!r} not found on the CV sheet")

    def _parse_region(self, ws, flow, start_row, end_row, quarter_cols) -> list[ContractRow]:
        co = load_company_resolver()
        seg_map = {
            h: (spec["segment"], spec["terminator"]) for h, spec in self.cfg["segments"].items()
        }
        terminators = {t for (_s, t) in seg_map.values()}
        industry_total = self.cfg["industry_total"]
        skip_rows = set(self.cfg.get("skip_rows", []))

        rows: list[ContractRow] = []
        cur_seg = cur_term = None
        for r in range(start_row, end_row):
            label = ws.cell(row=r, column=1).value
            if not isinstance(label, str) or not label.strip():
                continue
            label = label.strip()

            if label in seg_map:
                cur_seg, cur_term = seg_map[label]
                continue
            if label == industry_total:
                rows += self._emit(
                    ws,
                    r,
                    flow,
                    None,
                    INDUSTRY_TOTAL_CANONICAL,
                    label,
                    quarter_cols,
                    record_industry=True,
                )
                cur_seg = cur_term = None
                continue
            if cur_term is not None and label == cur_term:
                self._record_segment_total(ws, r, flow, cur_seg, quarter_cols)
                cur_seg = cur_term = None
                continue
            if label in terminators or label in skip_rows:
                continue  # a closed segment's terminator, or an intermediate subtotal
            if cur_seg is None:
                self.warnings.append(
                    f"CV/{flow}: skipped label outside a segment at row {r}: {label!r}"
                )
                continue
            try:
                canonical = co.resolve(label)
            except UnknownCompanyError:
                raise
            rows += self._emit(ws, r, flow, cur_seg, canonical, label, quarter_cols)
        return rows

    def _emit(
        self, ws, row, flow, segment, canonical, raw, quarter_cols, record_industry=False
    ) -> list[ContractRow]:
        out: list[ContractRow] = []
        for col, fq in quarter_cols.items():
            v = ws.cell(row=row, column=col).value
            if v is None or (isinstance(v, str) and v.strip().upper() in ("NA", "N/A", "-", "")):
                continue  # blank / "NA" = not reported (absence), not zero
            if not isinstance(v, (int, float)):
                self.warnings.append(f"non-numeric r{row}c{col} ({raw}): {v!r}")
                continue
            d = fiscal_quarter_start_date(fq)
            out.append(self._row(d, fq, flow, segment, canonical, raw, float(v)))
            if record_industry:
                self.reconciliation.industry_totals.setdefault(flow, {})[d] = float(v)
        return out

    def _record_segment_total(self, ws, row, flow, segment, quarter_cols) -> None:
        target = self.reconciliation.segment_totals.setdefault((flow, segment or ""), {})
        for col, fq in quarter_cols.items():
            v = ws.cell(row=row, column=col).value
            d = fiscal_quarter_start_date(fq)
            target[d] = float(v) if isinstance(v, (int, float)) else None

    def _row(self, d, fq, flow, segment, canonical, raw, value) -> ContractRow:
        return ContractRow(
            period_date=d,
            period_type="quarter",
            fiscal_year=fq[2:],
            fiscal_quarter=fq,
            category=self.category,
            segment=segment,
            sub_segment=None,
            company_canonical=canonical,
            company_raw=raw,
            flow=flow,
            powertrain="all",
            geography="IN",
            metric="units",
            value=value,
            unit="units",
            source=self.source_id,
            source_file=Path(self.source_file).name,
            source_period=self.source_period,
            native_frequency="quarter",
            calc_status="reported",
            revision=0,
            ingest_date=self.ingest_date,
            confidence="high",
            is_superseded=False,
            is_partial=False,
            periods_present=None,
            periods_expected=None,
        )
