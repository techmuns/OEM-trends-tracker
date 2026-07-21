"""VAHAN file adapter — parse a MANUALLY DOWNLOADED VAHAN dashboard export.

This is the free, sustainable path (see docs): a human downloads a report from the public
VAHAN dashboard (vahan.parivahan.gov.in/vahan4dashboard) in a browser — the dashboard blocks
cloud/CI runners, but a person on Indian internet reaches it fine — and drops the .xlsx here.
No scraping, no third-party API (both proven dead ends: the live site is unreachable from CI
and every free re-publisher is frozen years in the past).

VAHAN measures REGISTRATIONS, a different basis from SIAM's wholesale dispatches. Every row
emitted carries source="VAHAN"; the one-source-per-table rule keeps these from ever being
mixed with SIAM figures. This is a NEW, SEPARATE series — never an edit of the SIAM store.

Export shape (identical across the dashboard's Y-axes — Vehicle Class / Maker / Fuel):
    row 1: "<Dimension> Month Wise Data  For All State (<YEAR>)"
    row 2: S No | <Dimension> | Month Wise | ... | TOTAL
    row 4: '' | '' | JAN | FEB | ... | <last month> | ''      (month header)
    row 5+: <s.no> | <label> | <jan> | ... | <total>
Numbers use Indian grouping ("5,09,269"); "0" is a real reported zero (never null); a blank
cell is missing (never 0).

v1 scope: the **Vehicle Class** export -> per-category monthly registration totals
(2W/PV/3W/CV), via pipeline/dictionaries/vahan_classes.yaml. Maker- and Fuel-dimension files
parse too (see parse_report) but are not emitted as contract rows yet: both are all-category
in this export and can't be placed under the frozen per-category contract without a
category-filtered download. They are surfaced via ``report`` for inspection.
"""

from __future__ import annotations

import datetime as dt
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path

import openpyxl
import yaml

from pipeline.adapters.base import RawPayload, SourceAdapter, ValidationResult
from pipeline.aggregate.periods import fiscal_quarter_of, fiscal_year_of
from pipeline.contract.constants import INDUSTRY_TOTAL_CANONICAL
from pipeline.contract.models import ContractRow

_IST = dt.timezone(dt.timedelta(hours=5, minutes=30))
_DICT_PATH = Path(__file__).resolve().parents[1] / "dictionaries" / "vahan_classes.yaml"

_MONTHS = {
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}
# Row-2 header words that are NOT the dimension/label column.
_NON_LABEL_HEADERS = {"s no", "sno", "month wise", "total", ""}


class VahanFileError(RuntimeError):
    """The VAHAN export could not be parsed as expected."""


def _norm(s: object) -> str:
    """Collapse VAHAN's nbsp-padded labels to a clean, match-ready string."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).replace("\xa0", " ")).strip()


@lru_cache(maxsize=1)
def load_class_map(path: str | None = None) -> dict[str, str]:
    """VAHAN vehicle-class (upper-cased, whitespace-collapsed) -> contract category."""
    p = Path(path) if path else _DICT_PATH
    data = yaml.safe_load(p.read_text(encoding="utf-8"))
    return {_norm(k).upper(): v for k, v in data["classes"].items()}


@dataclass
class VahanReport:
    """The raw, dimension-agnostic parse of one export before any category mapping."""

    dimension: str  # "Vehicle Class" | "Maker" | "Fuel" | ...
    year: int
    months: list[int]  # calendar month numbers present, in column order
    # label -> {month_num: value|None}. value None = blank/missing (never coerced to 0).
    rows: dict[str, dict[int, float | None]] = field(default_factory=dict)


def _parse_number(cell: object) -> float | None:
    """'5,09,269' -> 509269.0 ; '' / None -> None (missing) ; '0' -> 0.0 (real zero)."""
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return float(cell)
    s = str(cell).replace(",", "").strip()
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_report(path: str | Path) -> VahanReport:
    """Parse any VAHAN '<Dimension> Month Wise Data' export into a VahanReport."""
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        title = _norm(ws.cell(1, 1).value)
        m = re.match(r"(.+?)\s+Month\s*Wise\s+Data", title, re.IGNORECASE)
        if not m:
            raise VahanFileError(f"unrecognized VAHAN title {title!r}")
        dimension = m.group(1).strip()
        ym = re.search(r"\((\d{4})\)", title)
        if not ym:
            raise VahanFileError(f"no year in VAHAN title {title!r}")
        year = int(ym.group(1))

        # Locate the month-header row (the one carrying JAN/FEB/... cells) and label column.
        month_cols: dict[int, int] = {}  # col_index -> month_num
        header_row = None
        for r in range(2, 8):
            hits = {
                c: _MONTHS[_norm(ws.cell(r, c).value).upper()]
                for c in range(1, ws.max_column + 1)
                if _norm(ws.cell(r, c).value).upper() in _MONTHS
            }
            if hits:
                header_row, month_cols = r, hits
                break
        if not month_cols:
            raise VahanFileError("no month columns (JAN..DEC) found in export")

        # Label column = row-2 cell that is neither S No / Month Wise / TOTAL.
        label_col = 2
        for c in range(1, ws.max_column + 1):
            if _norm(ws.cell(2, c).value).lower() not in _NON_LABEL_HEADERS:
                label_col = c
                break

        months = [month_cols[c] for c in sorted(month_cols)]
        rows: dict[str, dict[int, float | None]] = {}
        for r in range(header_row + 1, ws.max_row + 1):
            label = _norm(ws.cell(r, label_col).value)
            if not label or label.lower() in _NON_LABEL_HEADERS:
                continue
            series = {month_cols[c]: _parse_number(ws.cell(r, c).value) for c in month_cols}
            rows[label] = series
        return VahanReport(dimension=dimension, year=year, months=months, rows=rows)
    finally:
        wb.close()


class VahanFileAdapter(SourceAdapter):
    source_id = "VAHAN"
    native_frequency = "month"

    def __init__(self, source_file: str | Path, ingest_date: datetime | None = None) -> None:
        self.source_file = str(source_file)
        self.ingest_date = ingest_date or datetime.now(_IST)
        self.warnings: list[str] = []
        self.report: VahanReport | None = None

    def fetch(self, period: str) -> RawPayload:
        path = Path(self.source_file)
        if not path.exists():
            raise VahanFileError(f"source file not found: {path}")
        return RawPayload(self.source_id, period, self.source_file, content=path)

    def parse(self, raw: RawPayload) -> list[ContractRow]:
        report = parse_report(raw.content)
        self.report = report
        dim = report.dimension.lower()
        if "class" not in dim:
            # Maker / Fuel exports are all-category in this VAHAN report and cannot be placed
            # under the per-category contract without a category-filtered download. Parsed for
            # inspection (self.report) but no contract rows are emitted — never guess a category.
            self.warnings.append(
                f"{report.dimension!r} export parsed but not emitted as rows "
                "(all-category; needs a category-filtered download to place under the contract)."
            )
            return []
        return self._emit_class_totals(report)

    def _emit_class_totals(self, report: VahanReport) -> list[ContractRow]:
        """Vehicle Class -> per-(category, month) INDUSTRY-TOTAL registration rows."""
        class_map = load_class_map()
        # (category, month) -> summed registrations. Only present (non-null) cells contribute;
        # a category-month with no reported class stays absent (never a fabricated 0).
        totals: dict[tuple[str, int], float] = defaultdict(float)
        seen: set[tuple[str, int]] = set()
        for label, series in report.rows.items():
            category = class_map.get(label.upper())
            if category is None:
                self.warnings.append(
                    f"unmapped vehicle class {label!r} — ignored (not 2W/PV/3W/CV)"
                )
                continue
            for month, value in series.items():
                if value is None:
                    continue
                totals[(category, month)] += value
                seen.add((category, month))

        rows: list[ContractRow] = []
        for (category, month), value in sorted(totals.items()):
            d = date(report.year, month, 1)
            rows.append(self._row(d, category, value))
        return rows

    def _row(self, d: date, category: str, value: float) -> ContractRow:
        return ContractRow(
            period_date=d,
            period_type="month",
            fiscal_year=fiscal_year_of(d),
            fiscal_quarter=fiscal_quarter_of(d),
            category=category,
            segment=None,
            sub_segment=None,
            company_canonical=INDUSTRY_TOTAL_CANONICAL,
            company_raw="VAHAN registrations (all makers)",
            flow="domestic",
            powertrain="all",
            geography="IN",
            metric="units",
            value=value,
            unit="units",
            source=self.source_id,
            source_file=Path(self.source_file).name,
            source_period=f"{d.year}",
            native_frequency="month",
            calc_status="reported",
            revision=0,
            ingest_date=self.ingest_date,
            confidence="high",
            is_superseded=False,
            is_partial=False,
            periods_present=None,
            periods_expected=None,
        )

    def validate(self, rows: list[ContractRow]) -> ValidationResult:
        errors: list[str] = []
        warnings = list(self.warnings)
        if not rows:
            errors.append("no VAHAN category rows parsed (is this a Vehicle Class export?)")
        for r in rows:
            if r.source.value != "VAHAN":
                errors.append(f"non-VAHAN source leaked into VAHAN rows: {r.source}")
                break
        warnings += [
            f"negative registrations {r.value} for {r.category.value} {r.period_date}"
            for r in rows
            if r.value is not None and r.value < 0
        ]
        return ValidationResult(ok=not errors, errors=errors, warnings=warnings)
