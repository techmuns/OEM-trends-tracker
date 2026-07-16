"""VAHAN adapter — manually-downloaded registrations export (file drop, Phase 2+).

VAHAN reports **registrations**, a DIFFERENT measurement basis from SIAM wholesale
dispatches. Rows carry ``source='VAHAN'`` and never mix with SIAM in one table or share.

This parses the ``reportTable.xlsx`` that the vahan4dashboard "Report" download produces.
Two report shapes are supported, auto-detected from the sheet's own headers (never by
column position — the exact columns are read by their header text):

  * **Maker Month Wise**  — one row per manufacturer, monthly columns + a YTD total column.
  * **Fuel Month Wise**   — one row per fuel type, monthly columns + a YTD total column.

Layout (both):
  row 1  title, e.g. "Maker Month Wise Data  For All State (2026)"  ← report type + YEAR
  row 2  'S No' | ' Maker '/' Fuel ' | 'Month Wise ' | … | ' TOTAL '
  row 4  '' | '' | JAN | FEB | … | (blank over the TOTAL column)     ← month headers
  row 5+ serial | entity name | monthly values (Indian-comma strings) | YTD total

Scope note (all-India, ALL categories): the download the user exports is not filtered to a
vehicle class, so it is the whole registered universe (2W + PV + 3W + CV + tractors + …).
Rows are therefore emitted with ``category='ALL'`` and surface as a SEPARATE tab; they are
never merged into the SIAM per-category views.

Maker resolution deviates from SIAM's exhaustive hard-fail on purpose: the all-vehicle
export lists ~1,900 makers (a long tail of tiny/importer entities). We map an allow-list of
significant makers via ``companies.yaml`` and sum the unmapped remainder into the existing
``Others`` residual (the reported-universe remainder). A maker we DO map but that resolves
ambiguously still fails at dictionary-load time. VAHAN has NO exports and NO production, so
every row is ``flow='domestic'``.
"""

from __future__ import annotations

import datetime as dt
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl

from pipeline.adapters.base import RawPayload, SourceAdapter, ValidationResult
from pipeline.aggregate.periods import fiscal_quarter_of, fiscal_year_of
from pipeline.contract.constants import INDUSTRY_TOTAL_CANONICAL
from pipeline.contract.models import ContractRow
from pipeline.dictionaries.loader import UnknownCompanyError, load_company_resolver

SHEET = "reportTable"
_IST = dt.timezone(dt.timedelta(hours=5, minutes=30))
OTHERS_CANONICAL = "Others"

_MONTHS = {
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}

# Battery-electric fuels only. Hybrids (STRONG HYBRID EV, *…/HYBRID, PLUG-IN HYBRID EV) run
# an engine and are NOT counted as EV; hydrogen fuel-cell is a separate zero-emission class,
# also excluded. This mirrors the headline VAHAN "EV penetration" definition.
_EV_FUELS = {"PURE EV", "ELECTRIC(BOV)"}


class VahanFileParseError(RuntimeError):
    """The VAHAN reportTable workbook could not be parsed."""


def _clean(s: object) -> str:
    # VAHAN pads header/label cells with non-breaking spaces (\xa0). Collapse to a plain,
    # single-spaced, stripped string for matching.
    return re.sub(r"\s+", " ", str(s).replace("\xa0", " ")).strip()


def _num(v: object) -> float | None:
    """Indian-grouped registration count -> float. Blank/'-' = absence (None); '0' = zero."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s in ("", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


class VahanFileAdapter(SourceAdapter):
    source_id = "VAHAN"
    native_frequency = "month"

    def __init__(
        self,
        source_file: str | Path | list[str | Path],
        ingest_date: datetime | None = None,
        category: str = "ALL",
        source_period: str | None = None,
    ) -> None:
        # One VAHAN dataset can span two complementary exports (maker-wise + fuel-wise). They
        # are parsed and gated together — never as separate baselines — so a small fuel file
        # can't make the following maker file look like a 1000%+ row explosion.
        files = source_file if isinstance(source_file, (list, tuple)) else [source_file]
        self.source_files = [str(f) for f in files]
        self.source_file = self.source_files[0]  # first, for messages
        self.ingest_date = ingest_date or datetime(2026, 7, 15, 10, 0, tzinfo=_IST)
        self.category = category
        self.source_period = source_period
        self.modes: list[str] = []  # per-file "maker"/"fuel", set during parse
        self.warnings: list[str] = []

    # --- SourceAdapter interface -------------------------------------------------------

    def fetch(self, period: str) -> RawPayload:
        for f in self.source_files:
            if not Path(f).exists():
                raise VahanFileParseError(f"source file not found: {f}")
        return RawPayload(self.source_id, period, self.source_file, content=list(self.source_files))

    def parse(self, raw: RawPayload) -> list[ContractRow]:
        rows: list[ContractRow] = []
        self.modes = []
        for src in self.source_files:
            wb = openpyxl.load_workbook(src, data_only=True)
            try:
                if SHEET not in wb.sheetnames:
                    raise VahanFileParseError(
                        f"{Path(src).name}: sheet {SHEET!r} not found — not a VAHAN reportTable export"
                    )
                ws = wb[SHEET]
                mode, year = self._detect_mode_and_year(ws)
                self.modes.append(mode)
                month_row, month_cols = self._find_month_header(ws)
                entity_col = self._find_entity_col(ws, month_row, month_cols)
                data = self._read_table(ws, month_row, entity_col, month_cols)
                if not data:
                    raise VahanFileParseError(
                        f"{Path(src).name}: no data rows under the month header"
                    )
                sf = Path(src).name
                rows += (
                    self._emit_makers(data, year, sf)
                    if mode == "maker"
                    else self._emit_fuels(data, year, sf)
                )
            finally:
                wb.close()
        return rows

    def validate(self, rows: list[ContractRow]) -> ValidationResult:
        errors: list[str] = []
        if not rows:
            errors.append("no rows parsed from the VAHAN export")
        warnings = list(self.warnings)
        warnings += [
            f"negative registration {r.value} for {r.company_canonical} {r.period_date}"
            for r in rows
            if r.value is not None and r.value < 0
        ]
        return ValidationResult(ok=not errors, errors=errors, warnings=warnings)

    # --- detection / layout (header-driven, never positional) --------------------------

    def _detect_mode_and_year(self, ws) -> tuple[str, int]:
        # Report type + year both live in the title (row 1) and are corroborated by the
        # entity header (row 2). "Maker …" -> maker-wise; "Fuel …" -> fuel-wise.
        blob = " ".join(
            _clean(ws.cell(r, c).value)
            for r in range(1, 4)
            for c in range(1, min(ws.max_column, 4) + 1)
            if ws.cell(r, c).value is not None
        ).lower()
        mode = "maker" if "maker" in blob else "fuel" if "fuel" in blob else None
        if mode is None:
            raise VahanFileParseError(
                "could not tell maker-wise from fuel-wise (no 'Maker'/'Fuel' header)"
            )
        m = re.search(r"\b(20\d{2})\b", blob)
        if not m:
            raise VahanFileParseError("could not find a 4-digit year in the report title")
        return mode, int(m.group(1))

    def _find_month_header(self, ws) -> tuple[int, dict[int, int]]:
        """The row carrying JAN/FEB/… cells, and {column -> month number}. The YTD 'TOTAL'
        column has no month label there, so it is naturally excluded."""
        best_row, best_cols = 0, {}
        for r in range(1, 9):
            cols = {
                c: _MONTHS[_clean(v).upper()]
                for c in range(1, ws.max_column + 1)
                if (v := ws.cell(r, c).value) is not None and _clean(v).upper() in _MONTHS
            }
            if len(cols) > len(best_cols):
                best_row, best_cols = r, cols
        if not best_cols:
            raise VahanFileParseError("no month header row (JAN..DEC) found")
        return best_row, best_cols

    def _find_entity_col(self, ws, month_row: int, month_cols: dict[int, int]) -> int:
        # The entity column is the pre-month column whose DATA cells are text (fuel/maker
        # names), not numbers — 'S No' holds serials, the name column holds strings. Header
        # text is unreliable here: the row-1 title ("Maker/Fuel Month Wise …") sits in col 1
        # and would otherwise steal the match.
        first_month = min(month_cols)
        best_col, best_texty = max(1, first_month - 1), -1
        for c in range(1, first_month):
            texty = sum(
                1
                for r in range(month_row + 1, min(ws.max_row, month_row + 40) + 1)
                if (v := ws.cell(r, c).value) is not None and _clean(v) != "" and _num(v) is None
            )
            if texty > best_texty:
                best_col, best_texty = c, texty
        return best_col

    def _read_table(
        self, ws, month_row: int, entity_col: int, month_cols: dict[int, int]
    ) -> dict[str, dict[int, float | None]]:
        data: dict[str, dict[int, float | None]] = {}
        for r in range(month_row + 1, ws.max_row + 1):
            name = ws.cell(r, entity_col).value
            if name is None or _clean(name) == "":
                continue
            per = {month: _num(ws.cell(r, col).value) for col, month in month_cols.items()}
            data[_clean(name)] = per
        return data

    # --- emit ContractRows -------------------------------------------------------------

    def _emit_makers(
        self, data: dict[str, dict[int, float | None]], year: int, source_file: str
    ) -> list[ContractRow]:
        resolver = load_company_resolver()
        mapped: dict[str, dict[int, float]] = defaultdict(lambda: defaultdict(float))
        others: dict[int, float] = defaultdict(float)
        total: dict[int, float] = defaultdict(float)
        for raw, per in data.items():
            try:
                canonical = resolver.resolve(raw)
            except UnknownCompanyError:
                canonical = None  # long tail -> Others residual (documented VAHAN policy)
            for month, v in per.items():
                if v is None:
                    continue
                total[month] += v
                if canonical is None:
                    others[month] += v
                else:
                    mapped[canonical][month] += v

        rows: list[ContractRow] = []
        for canonical, per_m in mapped.items():
            for month, v in per_m.items():
                rows.append(self._row(year, month, canonical, canonical, v, "all", source_file))
        for month, v in others.items():
            rows.append(
                self._row(
                    year, month, OTHERS_CANONICAL, "(VAHAN unmapped makers)", v, "all", source_file
                )
            )
        # industry total = every registered maker (mapped + Others). For a complete all-India
        # registration census this sum IS the reported universe (every registration is
        # maker-attributed), so it is the honest denominator — not a lossy maker-sum.
        for month, v in total.items():
            rows.append(
                self._row(
                    year,
                    month,
                    INDUSTRY_TOTAL_CANONICAL,
                    "(VAHAN all-maker total)",
                    v,
                    "all",
                    source_file,
                )
            )
        return rows

    def _emit_fuels(
        self, data: dict[str, dict[int, float | None]], year: int, source_file: str
    ) -> list[ContractRow]:
        total: dict[int, float] = defaultdict(float)
        ev: dict[int, float] = defaultdict(float)
        for raw, per in data.items():
            is_ev = _clean(raw).upper() in _EV_FUELS
            for month, v in per.items():
                if v is None:
                    continue
                total[month] += v
                if is_ev:
                    ev[month] += v
        rows: list[ContractRow] = []
        for month, v in total.items():
            rows.append(
                self._row(
                    year,
                    month,
                    INDUSTRY_TOTAL_CANONICAL,
                    "(VAHAN all-fuel total)",
                    v,
                    "all",
                    source_file,
                )
            )
        for month, v in ev.items():
            rows.append(
                self._row(
                    year,
                    month,
                    INDUSTRY_TOTAL_CANONICAL,
                    "(VAHAN electric fuels)",
                    v,
                    "ev",
                    source_file,
                )
            )
        return rows

    def _row(
        self,
        year: int,
        month: int,
        canonical: str,
        raw: str,
        value: float,
        powertrain: str,
        source_file: str,
    ) -> ContractRow:
        d = date(year, month, 1)
        return ContractRow(
            period_date=d,
            period_type="month",
            fiscal_year=fiscal_year_of(d),
            fiscal_quarter=fiscal_quarter_of(d),
            category=self.category,  # "ALL" — all-India, all-vehicle registrations
            segment=None,
            sub_segment=None,
            company_canonical=canonical,
            company_raw=raw,
            flow="domestic",  # registrations; VAHAN has no export/production basis
            powertrain=powertrain,
            geography="IN",
            metric="units",
            value=value,
            unit="units",
            source=self.source_id,
            source_file=source_file,
            source_period=self.source_period or str(year),
            native_frequency="month",
            calc_status="reported",
            revision=0,
            ingest_date=self.ingest_date,
            confidence="high",
            is_superseded=False,
            is_partial=False,
            periods_present=None,
            periods_expected=None,
        )
