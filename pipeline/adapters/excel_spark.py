"""Excel 'Spark' summary-workbook adapter (File 1) — Phase 1 implementation.

Parses the sheet `OEM - Summary - 2W, PV, 3W`, **2W region only**, into contract rows.

Design decisions baked in (see the Phase 1 prompt):
- Load with `data_only=True`; both totals (`=+B35+B85`) and the date header row itself
  (`=+EOMONTH(EC2,0)+1`) are formulas. If cached values are missing we fail loudly.
- **Label-driven, never hardcoded rows.** We find `2W Domestic` / `Electric Two Wheelers`
  / `2W Exports` / `Total Domestic Two Wheelers` etc. and walk rows, so next month's file
  (which shifts row numbers) still parses.
- `all` includes EV; the EV block is a subset with `segment=null`. ICE is never parsed
  (it is derived `all − ev` downstream).
- `0` ≠ blank: a numeric `0` is a reported zero; a blank cell is absence (no row emitted).
- Per-segment `Total` rows are NOT ingested as companies — kept for reconciliation. The
  industry `Total Domestic/Exports Two Wheelers` rows ARE ingested as `reported` (they are
  the share denominator and define SIAM's reported universe).
- `source="SIAM"`: the figures are SIAM wholesale dispatches (the reported SIAM universe),
  even though this particular file is a manually maintained transcription. `source_file`
  preserves the workbook provenance. (Flagged for confirmation in the Phase 1 report.)
"""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import openpyxl

from pipeline.adapters.base import RawPayload, SourceAdapter, ValidationResult
from pipeline.aggregate.periods import fiscal_quarter_of, fiscal_year_of
from pipeline.contract.constants import INDUSTRY_TOTAL_CANONICAL
from pipeline.contract.models import ContractRow
from pipeline.dictionaries.loader import (
    UnknownCompanyError,
    load_company_resolver,
    load_segment_resolver,
)
from pipeline.validate.reconciliation import ReconciliationData, parse_file_quarter, series_key

SHEET = "OEM - Summary - 2W, PV, 3W"
CATEGORY = "2W"

DOMESTIC_START = "2W Domestic"
DOMESTIC_END = "Total Domestic Two Wheelers"
EXPORT_START = "2W Exports"
EXPORT_END = "Total Exports Two Wheelers"
REGION_END = "FOUR WHEELERS"
EV_HEADERS = {"Electric Two Wheelers", "Electric 2W"}
TOTAL_LABEL = "Total"

_IST = dt.timezone(dt.timedelta(hours=5, minutes=30))


class ExcelParseError(RuntimeError):
    """The workbook could not be parsed (e.g. missing cached values, missing labels)."""


@dataclass
class _HeaderMap:
    month_cols: dict[int, date]  # column index -> first-of-month date
    quarter_cols: dict[int, tuple[int, str]]  # column index -> (quarter_num, fiscal_year)


class ExcelSparkAdapter(SourceAdapter):
    source_id = "SIAM"
    native_frequency = "month"

    def __init__(
        self,
        source_file: str | Path,
        ingest_date: datetime | None = None,
        source_period: str = "2025-12",
    ) -> None:
        self.source_file = str(source_file)
        self.ingest_date = ingest_date or datetime(2026, 7, 15, 9, 0, tzinfo=_IST)
        self.source_period = source_period
        self.reconciliation = ReconciliationData()
        self.warnings: list[str] = []

    # --- SourceAdapter interface -------------------------------------------------------

    def fetch(self, period: str) -> RawPayload:
        path = Path(self.source_file)
        if not path.exists():
            raise ExcelParseError(f"source file not found: {path}")
        return RawPayload(
            source_id=self.source_id,
            source_period=period,
            source_file=self.source_file,
            content=path,
        )

    def parse(self, raw: RawPayload) -> list[ContractRow]:
        # NOT read_only: we do random cell access by (row, col); read_only mode makes that
        # O(n) per lookup. The sheet is small (~244x242), so a full load is fast.
        wb = openpyxl.load_workbook(raw.content, data_only=True)
        try:
            if SHEET not in wb.sheetnames:
                raise ExcelParseError(f"sheet {SHEET!r} not found in {self.source_file}")
            ws = wb[SHEET]
            header = self._read_header(ws)
            labels = self._find_labels(ws)
            self.reconciliation = ReconciliationData()
            rows: list[ContractRow] = []
            # A region spans from its start label to just before the next region. The EV
            # sub-block can appear AFTER the industry-total row (domestic) or before it
            # (exports), so we scan the whole region and emit the industry total inline
            # rather than stopping at it.
            rows += self._parse_region(
                ws,
                labels[DOMESTIC_START],
                labels[EXPORT_START] - 1,
                DOMESTIC_END,
                "domestic",
                header,
            )
            rows += self._parse_region(
                ws, labels[EXPORT_START], labels[REGION_END] - 1, EXPORT_END, "export", header
            )
            return rows
        finally:
            wb.close()

    def validate(self, rows: list[ContractRow]) -> ValidationResult:
        # Adapter-level self-check. Negatives are NOT errors — SIAM posts legitimate
        # negative monthly adjustments; the value_range_sanity gate flags them. Only a
        # structurally broken parse (no rows) is a hard error here.
        errors: list[str] = []
        warnings = list(self.warnings)
        if not rows:
            errors.append("no rows parsed")
        neg = [r for r in rows if r.value is not None and r.value < 0]
        warnings += [f"negative value {r.value} for {series_key(r)} {r.period_date}" for r in neg]
        return ValidationResult(ok=not errors, errors=errors, warnings=warnings)

    # --- helpers -----------------------------------------------------------------------

    def _read_header(self, ws) -> _HeaderMap:
        month_cols: dict[int, date] = {}
        quarter_cols: dict[int, tuple[int, str]] = {}
        for col in range(2, ws.max_column + 1):
            v = ws.cell(row=2, column=col).value
            if isinstance(v, (dt.datetime, dt.date)):
                month_cols[col] = date(v.year, v.month, 1)
            elif isinstance(v, str) and (qk := parse_file_quarter(v)) is not None:
                quarter_cols[col] = qk
        if not month_cols:
            raise ExcelParseError(
                "no monthly date columns found in header row 2 — cached values may be "
                "missing (file saved by a non-Excel tool). Refusing to guess."
            )
        return _HeaderMap(month_cols=month_cols, quarter_cols=quarter_cols)

    def _find_labels(self, ws) -> dict[str, int]:
        wanted = {DOMESTIC_START, DOMESTIC_END, EXPORT_START, EXPORT_END, REGION_END}
        found: dict[str, int] = {}
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if isinstance(a, str) and a.strip() in wanted and a.strip() not in found:
                found[a.strip()] = r
        missing = wanted - set(found)
        if missing:
            raise ExcelParseError(f"missing expected block labels: {sorted(missing)}")
        return found

    def _parse_region(
        self,
        ws,
        start_row: int,
        scan_end: int,
        industry_total_label: str,
        flow: str,
        header: _HeaderMap,
    ) -> list[ContractRow]:
        seg_resolver = load_segment_resolver()
        co_resolver = load_company_resolver()
        rows: list[ContractRow] = []

        current_segment: str | None = None
        current_powertrain = "all"

        for r in range(start_row + 1, scan_end + 1):
            label = ws.cell(row=r, column=1).value
            if not isinstance(label, str) or not label.strip():
                continue
            label = label.strip()

            if label == industry_total_label:  # the industry Total row -> ingest as reported
                rows += self._emit_series(
                    ws,
                    r,
                    flow,
                    powertrain="all",
                    segment=None,
                    canonical=INDUSTRY_TOTAL_CANONICAL,
                    raw=label,
                    calc_status="reported",
                    header=header,
                    record_industry=True,
                )
                current_segment = None
                current_powertrain = "all"
                continue

            if label in EV_HEADERS:
                current_powertrain = "ev"
                current_segment = None
                continue
            if seg_resolver.is_segment_header(CATEGORY, label):
                current_segment = seg_resolver.resolve(CATEGORY, label)
                current_powertrain = "all"
                continue
            if label == TOTAL_LABEL:
                # per-segment / EV block total -> reconciliation only, not a company
                self._record_block_total(ws, r, flow, current_powertrain, current_segment, header)
                current_segment = None
                continue

            # otherwise: a company row
            try:
                canonical = co_resolver.resolve(label)
            except UnknownCompanyError:
                raise  # hard fail — surfaces to the ingest
            rows += self._emit_series(
                ws,
                r,
                flow,
                powertrain=current_powertrain,
                segment=(current_segment if current_powertrain == "all" else None),
                canonical=canonical,
                raw=label,
                calc_status="reported",
                header=header,
            )
        return rows

    def _emit_series(
        self,
        ws,
        row: int,
        flow: str,
        powertrain: str,
        segment: str | None,
        canonical: str,
        raw: str,
        calc_status: str,
        header: _HeaderMap,
        record_industry: bool = False,
    ) -> list[ContractRow]:
        out: list[ContractRow] = []
        for col, d in header.month_cols.items():
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue  # blank = absence, NOT zero. Never emit a fabricated row.
            if not isinstance(v, (int, float)):
                self.warnings.append(f"non-numeric cell at r{row}c{col} ({raw}): {v!r}")
                continue
            out.append(
                self._row(d, flow, powertrain, segment, canonical, raw, float(v), calc_status)
            )
            if record_industry:
                self.reconciliation.industry_totals.setdefault(flow, {})[d] = float(v)
        # capture reported quarterly columns for this series (reconciliation (c))
        skey = (flow, powertrain, segment, canonical)
        qmap = self.reconciliation.reported_quarters.setdefault(skey, {})
        for col, qk in header.quarter_cols.items():
            v = ws.cell(row=row, column=col).value
            if isinstance(v, (int, float)):
                qmap[qk] = float(v)
        return out

    def _record_block_total(
        self, ws, row: int, flow: str, powertrain: str, segment: str | None, header: _HeaderMap
    ) -> None:
        target = (
            self.reconciliation.ev_totals.setdefault(flow, {})
            if powertrain == "ev"
            else self.reconciliation.segment_totals.setdefault((flow, segment or ""), {})
        )
        for col, d in header.month_cols.items():
            v = ws.cell(row=row, column=col).value
            target[d] = float(v) if isinstance(v, (int, float)) else None

    def _row(
        self,
        d: date,
        flow: str,
        powertrain: str,
        segment: str | None,
        canonical: str,
        raw: str,
        value: float,
        calc_status: str,
    ) -> ContractRow:
        return ContractRow(
            period_date=d,
            period_type="month",
            fiscal_year=fiscal_year_of(d),
            fiscal_quarter=fiscal_quarter_of(d),
            category=CATEGORY,
            segment=segment,
            sub_segment=None,
            company_canonical=canonical,
            company_raw=raw,
            flow=flow,
            powertrain=powertrain,
            geography="IN",
            metric="units",
            value=value,
            unit="units",
            source=self.source_id,
            source_file=Path(self.source_file).name,
            source_period=self.source_period,
            native_frequency="month",
            calc_status=calc_status,
            revision=0,
            ingest_date=self.ingest_date,
            confidence="high",
            is_superseded=False,
            is_partial=False,
            periods_present=None,
            periods_expected=None,
        )
