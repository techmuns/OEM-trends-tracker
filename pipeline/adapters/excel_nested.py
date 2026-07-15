"""Config-driven adapter for File 1's nested-block categories (PV now; 3W later).

Unlike 2W (a dedicated Electric subset block), PV/3W have **no EV block** — EV-only makers
sit inline among ICE makers, so EV is NOT derivable. Every row here is `powertrain=all`.
The block structure (segment headers, per-segment terminators which differ — "PC Volumes",
"UV Volumes", "Total Vans" — and industry-total labels) is read from
`dictionaries/categories.yaml`, never hardcoded.

Same non-negotiables as 2W: `data_only=True`, label-driven, unresolved maker → hard fail,
`0` ≠ blank, reported industry totals ingested (never reconstructed by summing).
"""

from __future__ import annotations

import datetime as dt
from datetime import date, datetime
from pathlib import Path

import openpyxl

from pipeline.adapters.base import RawPayload, SourceAdapter, ValidationResult
from pipeline.aggregate.periods import fiscal_quarter_of, fiscal_year_of
from pipeline.contract.constants import INDUSTRY_TOTAL_CANONICAL
from pipeline.contract.models import ContractRow
from pipeline.dictionaries.loader import (
    UnknownCompanyError,
    load_categories,
    load_company_resolver,
)
from pipeline.validate.reconciliation import ReconciliationData, parse_file_quarter, series_key

_IST = dt.timezone(dt.timedelta(hours=5, minutes=30))


class NestedBlockParseError(RuntimeError):
    pass


class NestedBlockAdapter(SourceAdapter):
    source_id = "SIAM"
    native_frequency = "month"

    def __init__(
        self,
        category: str,
        source_file: str | Path,
        ingest_date: datetime | None = None,
        source_period: str = "2025-12",
    ) -> None:
        self.category = category
        self.cfg = load_categories()[category]
        self.source_id = self.cfg.get("source", "SIAM")
        self.source_file = str(source_file)
        self.ingest_date = ingest_date or datetime(2026, 7, 15, 9, 0, tzinfo=_IST)
        self.source_period = source_period
        self.reconciliation = ReconciliationData()
        self.warnings: list[str] = []

    def fetch(self, period: str) -> RawPayload:
        path = Path(self.source_file)
        if not path.exists():
            raise NestedBlockParseError(f"source file not found: {path}")
        return RawPayload(self.source_id, period, self.source_file, content=path)

    def parse(self, raw: RawPayload) -> list[ContractRow]:
        wb = openpyxl.load_workbook(raw.content, data_only=True)
        try:
            sheet = self.cfg["sheet"]
            if sheet not in wb.sheetnames:
                raise NestedBlockParseError(f"sheet {sheet!r} not found")
            ws = wb[sheet]
            month_cols, quarter_cols = self._read_header(ws)
            start_row = self._find_label(ws, self.cfg["region_start"])
            end_row = self._find_label(ws, self.cfg["region_end"])
            self.reconciliation = ReconciliationData()
            return self._parse(ws, start_row, end_row, month_cols, quarter_cols)
        finally:
            wb.close()

    def validate(self, rows: list[ContractRow]) -> ValidationResult:
        errors: list[str] = []
        warnings = list(self.warnings)
        if not rows:
            errors.append("no rows parsed")
        warnings += [
            f"negative {r.value} for {series_key(r)} {r.period_date}"
            for r in rows
            if r.value is not None and r.value < 0
        ]
        return ValidationResult(ok=not errors, errors=errors, warnings=warnings)

    # --- helpers -----------------------------------------------------------------------

    def _read_header(self, ws) -> tuple[dict[int, date], dict[int, tuple[int, str]]]:
        month_cols: dict[int, date] = {}
        quarter_cols: dict[int, tuple[int, str]] = {}
        for col in range(2, ws.max_column + 1):
            v = ws.cell(row=2, column=col).value
            if isinstance(v, (dt.datetime, dt.date)):
                month_cols[col] = date(v.year, v.month, 1)
            elif isinstance(v, str) and (qk := parse_file_quarter(v)) is not None:
                quarter_cols[col] = qk
        if not month_cols:
            raise NestedBlockParseError("no monthly date columns in header row 2 (cached values?)")
        return month_cols, quarter_cols

    def _find_label(self, ws, label: str) -> int:
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if isinstance(a, str) and a.strip() == label:
                return r
        raise NestedBlockParseError(f"expected label {label!r} not found")

    def _parse(self, ws, start_row, end_row, month_cols, quarter_cols) -> list[ContractRow]:
        co = load_company_resolver()
        # flatten config: header-label -> (flow, segment, terminator); total-label -> flow
        seg_map: dict[str, tuple[str, str, str]] = {}
        for flow, headers in self.cfg["segments"].items():
            for header, spec in headers.items():
                seg_map[header] = (flow, spec["segment"], spec["terminator"])
        totals = {lbl: flow for flow, lbl in self.cfg["industry_totals"].items()}
        terminators = {t for (_f, _s, t) in seg_map.values()}

        rows: list[ContractRow] = []
        cur_flow = cur_seg = cur_term = None
        for r in range(start_row, end_row):
            label = ws.cell(row=r, column=1).value
            if not isinstance(label, str) or not label.strip():
                continue
            label = label.strip()

            if label in seg_map:
                cur_flow, cur_seg, cur_term = seg_map[label]
                continue
            if label in totals:
                rows += self._emit(ws, r, totals[label], None, INDUSTRY_TOTAL_CANONICAL,
                                    label, month_cols, quarter_cols, record_industry=True)
                cur_flow = cur_seg = cur_term = None
                continue
            if cur_term is not None and label == cur_term:
                self._record_segment_total(ws, r, cur_flow, cur_seg, month_cols)
                cur_seg = cur_term = None
                continue
            if label in terminators:
                continue  # a terminator for a different (already-closed) segment; skip
            if cur_flow is None or cur_seg is None:
                self.warnings.append(f"PV: stray label outside a segment at row {r}: {label!r}")
                continue
            # company row
            try:
                canonical = co.resolve(label)
            except UnknownCompanyError:
                raise
            rows += self._emit(ws, r, cur_flow, cur_seg, canonical, label, month_cols, quarter_cols)
        return rows

    def _emit(self, ws, row, flow, segment, canonical, raw, month_cols, quarter_cols,
              record_industry=False) -> list[ContractRow]:
        out: list[ContractRow] = []
        for col, d in month_cols.items():
            v = ws.cell(row=row, column=col).value
            if v is None or (isinstance(v, str) and v.strip().upper() in ("NA", "N/A", "-", "")):
                continue  # blank / "NA" = not reported (absence), not zero
            if not isinstance(v, (int, float)):
                self.warnings.append(f"non-numeric r{row}c{col} ({raw}): {v!r}")
                continue
            out.append(self._row(d, flow, segment, canonical, raw, float(v)))
            if record_industry:
                self.reconciliation.industry_totals.setdefault(flow, {})[d] = float(v)
        skey = (flow, "all", segment, canonical)
        qmap = self.reconciliation.reported_quarters.setdefault(skey, {})
        for col, qk in quarter_cols.items():
            v = ws.cell(row=row, column=col).value
            if isinstance(v, (int, float)):
                qmap[qk] = float(v)
        return out

    def _record_segment_total(self, ws, row, flow, segment, month_cols) -> None:
        target = self.reconciliation.segment_totals.setdefault((flow, segment or ""), {})
        for col, d in month_cols.items():
            v = ws.cell(row=row, column=col).value
            target[d] = float(v) if isinstance(v, (int, float)) else None

    def _row(self, d, flow, segment, canonical, raw, value) -> ContractRow:
        return ContractRow(
            period_date=d, period_type="month", fiscal_year=fiscal_year_of(d),
            fiscal_quarter=fiscal_quarter_of(d), category=self.category, segment=segment,
            sub_segment=None, company_canonical=canonical, company_raw=raw, flow=flow,
            powertrain="all", geography="IN", metric="units", value=value, unit="units",
            source=self.source_id, source_file=Path(self.source_file).name,
            source_period=self.source_period, native_frequency="month", calc_status="reported",
            revision=0, ingest_date=self.ingest_date, confidence="high", is_superseded=False,
            is_partial=False, periods_present=None, periods_expected=None,
        )
