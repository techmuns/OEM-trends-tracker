"""File 2 adapter — monthly SIAM industry workbook (`Monthly_SIAM_Industry_Data_*.xlsx`).

Different shape from File 1: one sheet per category; the `Two Wheelers` sheet is
TRANSPOSED — months run down column 1, and each maker is a group of columns
(Production / Domestic sales / Exports sales / Total sales). Maker-level only: NO segment,
NO EV/powertrain split.

Decision (extend-only, confirmed with the user): File 1 owns history through Dec-2025;
this adapter emits ONLY the forward extension (period >= FILE2_EXTEND_FROM), at maker level
(`segment=null`, `powertrain=all`). It does not supersede File 1's richer rows. For the
overlap it exposes `seam_reference` so `source_seam_check` can validate the join.

Non-negotiables (same as File 1): data_only=True with loud failure on missing cached
values; label-driven (no hardcoded indices); unresolved maker -> hard fail; 0 != blank;
inclusive-dimension guard still applies (we never store `total`, only domestic/export/
production, which are disjoint measurements).
"""

from __future__ import annotations

import datetime as dt
from datetime import date, datetime
from pathlib import Path

import openpyxl

from pipeline.adapters.base import RawPayload, SourceAdapter, ValidationResult
from pipeline.aggregate.periods import fiscal_quarter_of, fiscal_year_of
from pipeline.contract.constants import (
    FILE2_EXTEND_FROM,
    INDUSTRY_TOTAL_CANONICAL,
)
from pipeline.contract.models import ContractRow
from pipeline.dictionaries.loader import UnknownCompanyError, load_company_resolver

SHEET = "Two Wheelers"
CATEGORY = "2W"
_IST = dt.timezone(dt.timedelta(hours=5, minutes=30))

# metric column suffix -> contract flow ("Total sales" is skipped: total ⊇ domestic+export)
METRIC_FLOW = {
    "Production": "production",
    "Domestic sales": "domestic",
    "Exports sales": "export",
    "Total sales": None,
}
_SUFFIXES = tuple(METRIC_FLOW)  # for splitting "<maker> <metric>" combined headers

# Only compare the seam over a bounded recent overlap window (older years contain defunct
# makers absent from the dictionary; we neither ingest nor compare those).
SEAM_WINDOW_START = date(2024, 1, 1)
INDUSTRY_LABEL = "Industry"


class SiamMonthlyParseError(RuntimeError):
    """The monthly SIAM workbook could not be parsed."""


class SiamMonthlyAdapter(SourceAdapter):
    source_id = "SIAM"
    native_frequency = "month"

    def __init__(
        self,
        source_file: str | Path,
        ingest_date: datetime | None = None,
        source_period: str = "2026-05",
        extend_from: str = FILE2_EXTEND_FROM,
    ) -> None:
        self.source_file = str(source_file)
        self.ingest_date = ingest_date or datetime(2026, 7, 15, 9, 0, tzinfo=_IST)
        self.source_period = source_period
        self.extend_from = date.fromisoformat(extend_from)
        # seam_reference[(canonical, flow)] = {month: value} for overlap months (< extend_from)
        self.seam_reference: dict[tuple[str, str], dict[date, float]] = {}
        self.warnings: list[str] = []

    # --- SourceAdapter interface -------------------------------------------------------

    def fetch(self, period: str) -> RawPayload:
        path = Path(self.source_file)
        if not path.exists():
            raise SiamMonthlyParseError(f"source file not found: {path}")
        return RawPayload(self.source_id, period, self.source_file, content=path)

    def parse(self, raw: RawPayload) -> list[ContractRow]:
        wb = openpyxl.load_workbook(raw.content, data_only=True)
        try:
            if SHEET not in wb.sheetnames:
                raise SiamMonthlyParseError(f"sheet {SHEET!r} not found")
            ws = wb[SHEET]
            header_row = self._find_combined_header_row(ws)
            date_col, row_dates = self._find_date_column(ws, header_row)
            columns = self._map_columns(ws, header_row)  # col -> (maker_raw, flow)
            self.seam_reference = {}
            return self._emit(ws, columns, row_dates)
        finally:
            wb.close()

    def validate(self, rows: list[ContractRow]) -> ValidationResult:
        errors: list[str] = []
        warnings = list(self.warnings)
        if not rows:
            errors.append("no rows parsed (nothing at/after extend_from?)")
        warnings += [
            f"negative value {r.value} for {r.company_canonical}/{r.flow.value} {r.period_date}"
            for r in rows
            if r.value is not None and r.value < 0
        ]
        return ValidationResult(ok=not errors, errors=errors, warnings=warnings)

    # --- helpers -----------------------------------------------------------------------

    def _find_combined_header_row(self, ws) -> int:
        """The row whose cells read '<maker> <metric>' — detected, not hardcoded."""
        best_row, best_hits = None, 0
        for r in range(1, 9):
            hits = sum(
                1
                for c in range(1, ws.max_column + 1)
                if isinstance(ws.cell(r, c).value, str)
                and ws.cell(r, c).value.strip().endswith(_SUFFIXES)
            )
            if hits > best_hits:
                best_row, best_hits = r, hits
        if not best_row:
            raise SiamMonthlyParseError(
                "could not locate the maker/metric header row (missing cached values?)"
            )
        return best_row

    def _find_date_column(self, ws, header_row: int) -> tuple[int, dict[int, date]]:
        best_col, best = None, {}
        for c in range(1, min(ws.max_column, 4) + 1):
            dates = {
                r: date(v.year, v.month, 1)
                for r in range(header_row + 1, ws.max_row + 1)
                if isinstance((v := ws.cell(r, c).value), (dt.datetime, dt.date))
            }
            if len(dates) > len(best):
                best_col, best = c, dates
        if not best:
            raise SiamMonthlyParseError("no date column found (cached values absent?)")
        return best_col, best

    def _map_columns(self, ws, header_row: int) -> dict[int, tuple[str, str]]:
        cols: dict[int, tuple[str, str]] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if not isinstance(v, str):
                continue
            label = v.strip()
            for suf in _SUFFIXES:
                if label.endswith(suf):
                    flow = METRIC_FLOW[suf]
                    if flow is not None:  # skip "Total sales"
                        cols[c] = (label[: -len(suf)].strip(), flow)
                    break
        return cols

    def _resolve(self, maker_raw: str) -> str:
        if maker_raw == INDUSTRY_LABEL:
            return INDUSTRY_TOTAL_CANONICAL
        return load_company_resolver().resolve(maker_raw)

    def _emit(self, ws, columns, row_dates) -> list[ContractRow]:
        # group the metric columns by maker
        maker_cols: dict[str, dict[str, int]] = {}
        for col, (maker_raw, flow) in columns.items():
            maker_cols.setdefault(maker_raw, {})[flow] = col

        ingest_dates = {r: d for r, d in row_dates.items() if d >= self.extend_from}
        seam_dates = {
            r: d for r, d in row_dates.items() if SEAM_WINDOW_START <= d < self.extend_from
        }

        rows: list[ContractRow] = []
        for maker_raw, flowcols in maker_cols.items():
            # "active" = any nonzero value anywhere in the ingest window. A maker that is
            # zero across the ENTIRE window is inactive (defunct/not-in-2W) -> omit it: that
            # is absence, not a fabricated zero. Makers with mixed 0/nonzero keep their 0s.
            active = any(
                isinstance((v := ws.cell(r, c).value), (int, float)) and v != 0
                for c in flowcols.values()
                for r in ingest_dates
            )
            try:
                canonical = self._resolve(maker_raw)
            except UnknownCompanyError:
                if active:
                    raise  # an active ingest maker we cannot map -> hard fail
                self.warnings.append(f"skipped unknown inactive maker {maker_raw!r}")
                continue
            if active:
                for flow, c in flowcols.items():
                    for r, d in ingest_dates.items():
                        v = ws.cell(r, c).value
                        if v is None:
                            continue  # blank = absence, not zero
                        if not isinstance(v, (int, float)):
                            self.warnings.append(f"non-numeric {maker_raw} r{r}c{c}: {v!r}")
                            continue
                        rows.append(self._row(d, flow, canonical, maker_raw, float(v)))
            # seam reference (domestic+export) for dictionary makers, comparable to File 1
            for flow in ("domestic", "export"):
                c = flowcols.get(flow)
                if c is None:
                    continue
                for r, d in seam_dates.items():
                    v = ws.cell(r, c).value
                    if isinstance(v, (int, float)):
                        self.seam_reference.setdefault((canonical, flow), {})[d] = float(v)
        return rows

    def _row(self, d: date, flow: str, canonical: str, raw: str, value: float) -> ContractRow:
        return ContractRow(
            period_date=d,
            period_type="month",
            fiscal_year=fiscal_year_of(d),
            fiscal_quarter=fiscal_quarter_of(d),
            category=CATEGORY,
            segment=None,  # File 2 has no segment split
            sub_segment=None,
            company_canonical=canonical,
            company_raw=raw,
            flow=flow,
            powertrain="all",  # no EV split in File 2
            geography="IN",
            metric="units",
            value=value,
            unit="units",
            source=self.source_id,
            source_file=Path(self.source_file).name,
            source_period=self.source_period,
            native_frequency="month",
            calc_status="reported",
            revision=0,
            ingest_date=self.ingest_date,
            confidence="high",
            is_superseded=False,
            is_partial=False,
            periods_present=None,
            periods_expected=None,
        )
